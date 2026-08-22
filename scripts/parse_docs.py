"""Parse the reference documents into two kinds of artifact.

  code tables -> data/reference/*.csv   joined in SQL, never embedded
  prose       -> data/reference/chunks.jsonl   embedded by build_index.py

Why the split: a code table is an exact lookup. "What is occupancy code 105?"
must return code 105, not the nearest neighbour in embedding space. Fuzzy
retrieval over an exact key is a correctness bug, so those go to SQL.

Run from the repo root:  python scripts/parse_docs.py
"""

import json
import re
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook

REF = Path("data/reference")
chunks: list[dict] = []


def add(text: str, source: str, section: str) -> None:
    text = re.sub(r"[ \t]+", " ", text).strip()
    if len(text) > 40:
        chunks.append({"text": text, "source": source, "section": section})


# ---------------------------------------------------------------------------
# Code tables -> CSV for SQL
# ---------------------------------------------------------------------------

def parse_offense_codes() -> int:
    """rmsoffensecodes.xlsx: 576 rows, 425 distinct codes, 65 conflicting
    names. Deduplicate to one row per code or a join fans crime out by 70%."""
    ws = load_workbook(REF / "rmsoffensecodes.xlsx", read_only=True).worksheets[0]
    seen: dict[str, str] = {}
    for code, name in ws.iter_rows(min_row=2, values_only=True):
        if code is None:
            continue
        code = str(code).strip()
        name = (str(name).strip() if name else "")
        # keep the first spelling we see; they differ only by typos
        seen.setdefault(code, name)
    out = REF / "offense_codes.csv"
    with out.open("w") as f:
        f.write("offense_code,offense_name\n")
        for code, name in sorted(seen.items()):
            f.write(f'{code},"{name}"\n')
    return len(seen)


def parse_occupancy_codes() -> int:
    """propertyoccupancycodes.pdf is a 4-up layout, so extract_tables() finds
    nothing. Pull `NNN DESCRIPTION` pairs out of the raw text instead."""
    with pdfplumber.open(REF / "property_occupancy_codes.pdf") as pdf:
        text = "\n".join(p.extract_text() or "" for p in pdf.pages)
    pairs = re.findall(r"\b(\d{3})\s+([A-Z0-9][A-Z0-9 /&'\.\-]{2,40}?)(?=\s+\d{3}\s|\s*$)",
                       text, flags=re.MULTILINE)
    seen: dict[str, str] = {}
    for code, desc in pairs:
        seen.setdefault(code, desc.strip())
    out = REF / "occupancy_codes.csv"
    with out.open("w") as f:
        f.write("occupancy_code,description\n")
        for code, desc in sorted(seen.items()):
            f.write(f'{code},"{desc}"\n')
    return len(seen)


# ---------------------------------------------------------------------------
# Prose -> chunks for embedding
# ---------------------------------------------------------------------------

def parse_field_tables(filename: str, source: str) -> None:
    """311_data_dictionary.pdf and property_fy2026_data_key.pdf extract as
    clean 'name | description | type | example' tables. One chunk per field.

    WARNING: the 311 dictionary documents a DIFFERENT schema than the CSV in
    this repo - only 10 of its 28 column names actually exist. These chunks
    are for definitional questions only. Never feed them to SQL generation;
    use data/DATA_DICTIONARY.md for that.
    """
    with pdfplumber.open(REF / filename) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                for row in table:
                    if not row or not row[0]:
                        continue
                    name = row[0].replace("\n", " ").strip()
                    if name.lower() in ("column name", "field name", ""):
                        continue
                    body = " ".join(
                        c.replace("\n", " ").strip() for c in row[1:] if c
                    )
                    if body:
                        add(f"{name}: {body}", source, name)


def parse_crm_value_codes() -> None:
    """311_crm_value_codes.pdf is prose with bullet lists, not a table.
    Split on the field headings so each closure code keeps its meaning."""
    with pdfplumber.open(REF / "311_crm_value_codes.pdf") as pdf:
        text = "\n".join(p.extract_text() or "" for p in pdf.pages)
    text = text.replace("\xad", "-").replace("○", "-")
    # headings look like  CLOSURE_REASON - ...  or  CASE_TITLE - ...
    parts = re.split(r"\n(?=[A-Z][A-Z_]{3,}\s*-)", text)
    for part in parts:
        head = part.split("-", 1)[0].strip()[:60] or "general"
        for line in part.split("\n"):
            add(line, "311_crm_value_codes.pdf", head)


def parse_crime_fields() -> None:
    ws = load_workbook(REF / "crime_field_explanation.xlsx", read_only=True).worksheets[0]
    for name, desc in ws.iter_rows(min_row=2, values_only=True):
        if name and desc:
            add(f"{name}: {desc}", "crime_field_explanation.xlsx", str(name))


def main() -> None:
    n_off = parse_offense_codes()
    n_occ = parse_occupancy_codes()
    print(f"code tables -> SQL")
    print(f"  offense_codes.csv    {n_off:>5} codes (deduplicated)")
    print(f"  occupancy_codes.csv  {n_occ:>5} codes")

    parse_field_tables("311_data_dictionary.pdf", "311_data_dictionary.pdf")
    parse_field_tables("property_fy2026_data_key.pdf", "property_fy2026_data_key.pdf")
    parse_crm_value_codes()
    parse_crime_fields()

    out = REF / "chunks.jsonl"
    with out.open("w") as f:
        for c in chunks:
            f.write(json.dumps(c) + "\n")

    print(f"\nprose -> embeddings")
    print(f"  chunks.jsonl         {len(chunks):>5} chunks")
    by_source: dict[str, int] = {}
    for c in chunks:
        by_source[c["source"]] = by_source.get(c["source"], 0) + 1
    for s, n in sorted(by_source.items()):
        print(f"    {s:38} {n:>4}")


if __name__ == "__main__":
    main()
